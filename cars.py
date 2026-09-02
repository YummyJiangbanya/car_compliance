import os
import re
import sqlite3
import pandas as pd
import openpyxl
import streamlit as st

# ==================== 1. 页面配置 ====================
st.set_page_config(
    page_title="智能网联汽车与跨国数据合规检索平台 | Newsprint Edition",
    page_icon="📰", 
    layout="wide",
    initial_sidebar_state="collapsed"
)

# 初始化 Session State
if "nav_choice" not in st.session_state:
    st.session_state.nav_choice = "首页"
if "show_terms_page" not in st.session_state:
    st.session_state.show_terms_page = False
if "selected_case" not in st.session_state:
    st.session_state.selected_case = None

# ==================== 2. 全局 CSS 样式与 UI 设计系统 (Newsprint 风格) ====================
NEWSPRINT_CSS = """
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=JetBrains+Mono:wght@400;500&family=Playfair+Display:ital,wght@0,400;0,600;0,700;0,900;1,400&family=Lora:ital,wght@0,400;0,600;1,400&display=swap');
    :root {
        --bg-base: #F9F9F7;
        --bg-surface: #F9F9F7;
        --text-primary: #111111;
        --text-muted: #666666;
        --border-color: #111111;
        --accent-red: #CC0000;
        --divider-grey: #E5E5E0;
    }
    /* 页面基础背景 */
    [data-testid="stAppViewContainer"] {
        background-color: var(--bg-base) !important;
        background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='4' height='4' viewBox='0 0 4 4'%3E%3Cpath fill='%23111111' fill-opacity='0.04' d='M1 3h1v1H1V3zm2-2h1v1H3V1z'%3E%3C/path%3E%3C/svg%3E") !important;
        color: var(--text-primary) !important;
    }
    
    [data-testid="stHeader"] {
        background-color: transparent !important;
    }
    /* 全局字体定义 */
    html, body, p, label, li, .law-content {
        font-family: 'Lora', Georgia, serif;
        color: var(--text-primary);
    }
    /* 标题统合 */
    h1, h2, h3, h4 {
        font-family: 'Playfair Display', 'Times New Roman', serif !important;
        color: #111111 !important;
        font-weight: 900 !important;
        letter-spacing: -0.03em;
        border-bottom: 2px solid #111111;
        padding-bottom: 8px;
        margin-bottom: 20px;
    }
    /* 绝对零圆角与硬阴影悬浮效果 */
    .sharp-card, div[data-testid="stExpander"], .term-card, .timeline-card, .header-card {
        background-color: #F9F9F7 !important;
        border: 1px solid #111111 !important;
        border-radius: 0px !important;
        box-shadow: none !important;
        transition: all 150ms cubic-bezier(0, 0, 0.2, 1);
        padding: 24px;
        margin-bottom: 20px;
    }
    .hard-shadow-hover:hover {
        box-shadow: 4px 4px 0px 0px #111111 !important;
        transform: translate(-2px, -2px);
    }
    /* Expander 样式调整 */
    div[data-testid="stExpander"] { padding: 0 !important; }
    div[data-testid="stExpander"] summary {
        padding: 16px 20px;
        background-color: #F9F9F7;
        border: 1px solid #111111;
        border-radius: 0px !important;
    }
    div[data-testid="stExpander"] summary span[data-testid="stExpanderToggleIcon"] {
        font-size: 0px !important;
        color: transparent !important;
        width: 0px !important;
        display: none !important;
    }
    div[data-testid="stExpander"] summary:hover {
        background-color: #111111 !important;
        color: #F9F9F7 !important;
    }
    div[data-testid="stExpander"] summary p {
        font-family: 'Playfair Display', serif !important;
        font-weight: 700;
        color: #111111 !important;
        margin: 0 !important;
    }
    div[data-testid="stExpander"] summary:hover p {
        color: #F9F9F7 !important;
    }
    /* 隐藏默认侧边栏 */
    [data-testid="stSidebar"] {
        display: none !important;
    }
    /* 顶端导航链接 */
    .nav-tabs-container {
        display: flex;
        border: 1px solid #111111;
        background-color: #F9F9F7;
        margin-bottom: 25px;
    }
    .nav-tab-item {
        flex: 1;
        text-align: center;
        border-right: 1px solid #111111;
    }
    .nav-tab-item:last-child {
        border-right: none;
    }
    .nav-tab-item button {
        background-color: transparent !important;
        color: #111111 !important;
        border: none !important;
        border-radius: 0px !important;
        font-family: 'Inter', sans-serif !important;
        font-weight: 700 !important;
        font-size: 0.9rem !important;
        text-transform: uppercase;
        letter-spacing: 0.05em;
        box-shadow: none !important;
        width: 100% !important;
        padding: 12px 0px !important;
        margin: 0px !important;
        text-align: center;
        transition: all 100ms ease !important;
    }
    .nav-tab-item button:hover {
        background-color: #111111 !important;
        color: #F9F9F7 !important;
    }
    .nav-tab-active button {
        background-color: #111111 !important;
        color: #F9F9F7 !important;
    }
    /* 术语按钮样式 */
    .inline-term-btn button {
        background-color: #F9F9F7 !important;
        color: #111111 !important;
        border: 1px solid #111111 !important;
        border-radius: 0px !important;
        font-family: 'Inter', sans-serif !important;
        font-weight: 600 !important;
        font-size: 0.85rem !important;
        text-transform: uppercase;
        letter-spacing: 0.05em;
        box-shadow: none !important;
        padding: 8px 16px !important;
        transition: background-color 100ms ease, color 100ms ease !important;
        white-space: nowrap;
    }
    .inline-term-btn button:hover {
        background-color: #111111 !important;
        color: #F9F9F7 !important;
        border: 1px solid #111111 !important;
        box-shadow: 2px 2px 0px 0px #111111 !important;
    }
    /* 报纸风格标签 */
    .law-tag {
        display: inline-block;
        background-color: #111111;
        color: #F9F9F7;
        padding: 4px 10px;
        border-radius: 0px !important;
        font-family: 'JetBrains Mono', monospace;
        font-size: 0.75rem;
        text-transform: uppercase;
        letter-spacing: 0.05em;
        margin-bottom: 8px;
        margin-right: 6px;
        border: 1px solid #111111;
    }
    /* 条款排版 */
    .law-content {
        background-color: #ffffff;
        border: 1px solid #111111;
        border-left: 6px solid #111111 !important;
        padding: 20px;
        color: #111111;
        line-height: 1.8;
        font-size: 1rem;
        text-align: justify;
        white-space: pre-wrap;
        border-radius: 0px !important;
    }
    /* 术语卡片 */
    .term-card {
        border-left: 6px solid var(--accent-red) !important;
    }
    .term-source {
        font-family: 'JetBrains Mono', monospace;
        color: #666666;
        font-size: 0.8rem;
        margin-top: 15px;
        text-align: right;
    }
    /* 时间轴 */
    .timeline-container {
        position: relative;
        padding-left: 30px;
        margin: 30px 0;
        border-left: 2px solid #111111;
    }
    .timeline-item {
        position: relative;
        margin-bottom: 35px;
    }
    .timeline-node {
        position: absolute;
        left: -37px;
        top: 18px;
        width: 12px;
        height: 12px;
        border-radius: 0px !important;
        background-color: #F9F9F7;
        border: 3px solid #111111;
    }
    /* 输入框设计 */
    div[data-testid="stTextInput"] input {
        background-color: transparent !important;
        border: none !important;
        border-bottom: 2px solid #111111 !important;
        color: #111111 !important;
        font-family: 'JetBrains Mono', monospace !important;
        border-radius: 0px !important;
        box-shadow: none !important;
    }
    div[data-testid="stTextInput"] input:focus {
        background-color: #E5E5E0 !important;
    }
    div[data-testid="stTextInput"] label {
        font-family: 'Inter', sans-serif !important;
        font-weight: 700 !important;
        text-transform: uppercase;
        font-size: 0.8rem;
    }
    /* 报头元数据 */
    .newsprint-masthead {
        border-top: 3px solid #111111;
        border-bottom: 1px solid #111111;
        padding: 8px 0;
        margin-bottom: 24px;
        display: flex;
        justify-content: space-between;
        font-family: 'JetBrains Mono', monospace;
        font-size: 0.75rem;
        text-transform: uppercase;
        letter-spacing: 0.1em;
    }
</style>
"""
st.markdown(NEWSPRINT_CSS, unsafe_allow_html=True)
DB_FILE = "car_compliance.db"

# ==================== 3. 核心处理与数据库函数 ====================
def extract_sort_key(text):
    match_cn = re.search(r"第([零一二三四五六七八九十百0-9]+)条", text)
    if match_cn:
        num_str = match_cn.group(1)
        mapping = {"一":1, "二":2, "三":3, "四":4, "五":5, "六":6, "七":7, "八":8, "九":9, "十":10,
                   "十一":11, "十二":12, "十三":13, "十四":14, "十五":15, "十六":16, "十七":17, "十八":18, "十九":19, "二十":20,
                   "二十一":21, "二十二":22, "二十三":23, "二十四":24, "二十五":25, "二十六":26, "二十七":27, "二十八":28, "二十九":29, "三十":30,
                   "三十一":31, "三十二":32, "三十三":33, "三十四":34, "三十五":35, "三十六":36, "三十七":37, "三十八":38, "三十九":39, "四十":40}
        if num_str in mapping: return mapping[num_str]
        try: return int(num_str)
        except ValueError: pass
    match_en = re.search(r"Article\s+(\d+)", text, re.IGNORECASE)
    if match_en:
        try: return int(match_en.group(1))
        except ValueError: pass
    return 999

def get_clean_cell_text(cell):
    if cell.value is None or str(cell.value).strip() == "nan":
        return ""
    
    is_rich = hasattr(cell, 'value') and isinstance(cell.value, openpyxl.cell.rich_text.CellRichText)
    if is_rich:
        full_text = "".join([str(rt.text) for rt in cell.value])
    else:
        full_text = str(cell.value)
    
    return full_text.strip()

@st.cache_data
def init_database_from_excel():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    
    # 智能查找最新的 Excel 文件
    possible_names = [
        "合规平台条文整理（修改4.0）_4.xlsx",
        "合规平台条文整理（修改4.0）.xlsx",
        "合规平台条文整理（修改1.0）.xlsx",
        "合规平台条文整理 (1).xlsx"
    ]
    excel_path = None
    for fname in possible_names:
        p = os.path.join(current_dir, fname)
        if os.path.exists(p):
            excel_path = p
            break
    if not excel_path:
        for fname in os.listdir(current_dir):
            if fname.endswith(".xlsx") and "合规平台" in fname:
                excel_path = os.path.join(current_dir, fname)
                break

    if not excel_path or not os.path.exists(excel_path):
        return False

    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS compliance_laws (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            region TEXT,
            category TEXT,
            law_title TEXT,
            sub_cat_0 TEXT,
            sub_cat_1 TEXT,
            content TEXT,
            sort_order INTEGER
        )
    """)
    cursor.execute("DELETE FROM compliance_laws")

    wb = openpyxl.load_workbook(excel_path, data_only=False)
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    df_raw = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)

    categories_row = df_raw.iloc[0]
    titles_row = df_raw.iloc[1]
    col0_raw = df_raw.iloc[:, 0] if len(df_raw.columns) > 0 else pd.Series([""] * len(df_raw))
    col1_raw = df_raw.iloc[:, 1] if len(df_raw.columns) > 1 else pd.Series([""] * len(df_raw))
    col2_raw = df_raw.iloc[:, 2] if len(df_raw.columns) > 2 else pd.Series([""] * len(df_raw))

    laws_dict = {}
    laws_order = []
    all_law_columns = []

    # 遍历第 3 列及以后的所有列（法规列）
    for col_idx in range(3, len(df_raw.columns)):
        cat_raw = str(categories_row.iloc[col_idx]).strip()
        law_title = str(titles_row.iloc[col_idx]).strip()

        if not law_title or law_title == "nan":
            continue

        # 第一行精准拆分 司法辖区 与 合规模块
        if "-" in cat_raw:
            parts = cat_raw.split("-", 1)
            region = parts[0].strip()
            category = parts[1].strip()
        elif "—" in cat_raw:
            parts = cat_raw.split("—", 1)
            region = parts[0].strip()
            category = parts[1].strip()
        elif "–" in cat_raw:
            parts = cat_raw.split("–", 1)
            region = parts[0].strip()
            category = parts[1].strip()
        else:
            if "欧盟" in cat_raw:
                region = "欧盟"
                category = cat_raw.replace("欧盟", "").strip() or "通用模块"
            elif "美国" in cat_raw:
                region = "美国"
                category = cat_raw.replace("美国", "").strip() or "通用模块"
            elif "中国" in cat_raw:
                region = "中国"
                category = cat_raw.replace("中国", "").strip() or "通用模块"
            else:
                region = "中国"
                category = cat_raw if cat_raw and cat_raw != "nan" else "通用模块"

        all_law_columns.append((region, category, law_title))

        # 遍历全列的所有数据单元格（第 2 行开始）
        for row_idx in range(2, len(df_raw)):
            cell_obj = ws.cell(row=row_idx + 1, column=col_idx + 1)
            content_str = get_clean_cell_text(cell_obj)

            if content_str and content_str != "nan":
                # 拼接第 0、1、2 列作为适用场景
                s0 = str(col0_raw.iloc[row_idx]).strip()
                s1 = str(col1_raw.iloc[row_idx]).strip()
                s2 = str(col2_raw.iloc[row_idx]).strip() if len(df_raw.columns) > 2 else ""

                sc_parts = [x for x in [s0, s1, s2] if x and x != "nan"]
                scenario_text = " ➔ ".join(sc_parts) if sc_parts else "通用指引"

                sort_val = extract_sort_key(content_str)
                key = (region, category, law_title, content_str)

                # 识别重复法条：若同一个法规下的法条内容完全一样，合并条目并累计不同的适用场景
                if key not in laws_dict:
                    laws_dict[key] = {
                        "region": region,
                        "category": category,
                        "law_title": law_title,
                        "content": content_str,
                        "sort_order": sort_val,
                        "scenarios": []
                    }
                    laws_order.append(key)

                if scenario_text not in laws_dict[key]["scenarios"]:
                    laws_dict[key]["scenarios"].append(scenario_text)

    # 将合并后的法条数据写入 SQLite 数据库
    processed_laws = set()
    for key in laws_order:
        item = laws_dict[key]
        combined_scenarios = " | ".join(item["scenarios"])
        cursor.execute(
            "INSERT INTO compliance_laws (region, category, law_title, sub_cat_0, sub_cat_1, content, sort_order) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (item["region"], item["category"], item["law_title"], combined_scenarios, "", item["content"], item["sort_order"])
        )
        processed_laws.add((item["region"], item["category"], item["law_title"]))

    # 保留占位（应对暂无可检索条文的法规列）
    for (r, c, lt) in set(all_law_columns):
        if (r, c, lt) not in processed_laws:
            cursor.execute(
                "INSERT INTO compliance_laws (region, category, law_title, sub_cat_0, sub_cat_1, content, sort_order) VALUES (?, ?, ?, ?, ?, ?, ?)",
                (r, c, lt, "暂无分类", "暂无指引", "（该法规条文正在整理补充中，敬请期待...）", 999)
            )

    conn.commit()
    conn.close()
    return True

success_db = init_database_from_excel()

# ==================== 4. 顶端导航选项卡 ====================
nav_items = [
    ("首页", "首页"),
    ("法律库", "法律库"),
    ("出境全流程时间轴", "出境全流程时间轴"),
    ("案例库", "案例库"),
    ("关于我们", "关于我们")
]

top_cols = st.columns(5)
for idx, (label, choice_key) in enumerate(nav_items):
    with top_cols[idx]:
        is_active = (st.session_state.nav_choice == choice_key)
        active_style_class = "nav-tab-active" if is_active else ""
        st.markdown(f'<div class="nav-tab-item {active_style_class}" style="border:none; margin-bottom:15px;">', unsafe_allow_html=True)
        if st.button(label, key=f"nav_btn_{idx}", use_container_width=True):
            st.session_state.nav_choice = choice_key
            st.session_state.show_terms_page = False
            st.session_state.selected_case = None
            st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

# ==================== 5. 页面展示逻辑 ====================
if st.session_state.show_terms_page:
    st.markdown("<h2 style='text-align: center; border-bottom: 3px solid #111;'>📖 术语解释总结全库专栏</h2>", unsafe_allow_html=True)
    st.markdown("<p style='text-align: center; font-family: Lora, serif; color: #666666;'>展示完整的汽车数据及出境合规术语释义，还原现代纸媒专栏的严谨审慎与清晰结构。</p>", unsafe_allow_html=True)
    term_keyword = st.text_input("🔍 检索术语关键字 (如：个人信息、重要数据、GDPR...)", key="standalone_term_search", placeholder="在此输入关键字进行检索...")
    current_dir = os.path.dirname(os.path.abspath(__file__))
    term_excel_path = os.path.join(current_dir, "术语解释总结.xlsx")
    if os.path.exists(term_excel_path):
        try:
            wb = openpyxl.load_workbook(term_excel_path, data_only=True)
            ws = wb.active
            law_names = []
            for cell in ws[1]:
                law_names.append(str(cell.value).strip() if cell.value else "")
            terms_list = []
            for row in ws.iter_rows(min_row=2):
                for c_idx, cell in enumerate(row):
                    if cell.value:
                        cell_str = str(cell.value).strip()
                        if not cell_str or cell_str.lower() == "nan":
                            continue
                        source_law = law_names[c_idx] if c_idx < len(law_names) and law_names[c_idx] else "未知法规"
                        parts = cell_str.split(":", 1)
                        if len(parts) == 2:
                            term_name = parts[0].strip()
                            definition = parts[1].strip()
                        else:
                            term_name = cell_str
                            definition = cell_str
                        color = None
                        if cell.fill and cell.fill.start_color:
                            color_val = cell.fill.start_color.index
                            if color_val and str(color_val) != '00000000':
                                color = str(color_val)
                        terms_list.append({
                            "term_name": term_name,
                            "definition": definition,
                            "source": source_law,
                            "color": color,
                            "original_full": cell_str
                        })
            final_results = []
            if term_keyword:
                term_keyword_lower = term_keyword.lower()
                matched_colors = set()
                direct_match_indices = set()
                for i, t in enumerate(terms_list):
                    if term_keyword_lower in t["term_name"].lower():
                        direct_match_indices.add(i)
                        if t["color"]:
                            matched_colors.add(t["color"])
                for i, t in enumerate(terms_list):
                    if i in direct_match_indices or (t["color"] and t["color"] in matched_colors):
                        final_results.append(t)
                st.markdown(f"**检索到相关术语/条文共计：{len(final_results)} 条**")
            else:
                final_results = terms_list
                st.markdown(f"**当前库内完整术语/条文共计：{len(final_results)} 条**")
            for t_item in final_results:
                if t_item['term_name'] != t_item['definition']:
                    def_html = f"<b>{t_item['term_name']}：</b>" + t_item['definition']
                else:
                    def_html = t_item['original_full']
                source_text = f"（来源：《{t_item['source']}》）"
                st.markdown(
                    f"""
                    <div class="term-card hard-shadow-hover">
                        <div>{def_html}</div>
                        <div class="term-source">{source_text}</div>
                    </div>
                    """, 
                    unsafe_allow_html=True
                )
        except Exception as e:
            st.error(f"加载术语表异常: {e}")
    else:
        st.warning("未检测到 `术语解释总结.xlsx` 文件，请确认已上传至同一目录。")
else:
    if st.session_state.nav_choice == "首页":
        st.markdown(
            """
            <div class="newsprint-masthead">
                <span>VOL. I NO. 01</span>
                <span>AUTOMOTIVE DATA COMPLIANCE REVIEW</span>
                <span>GLOBAL EDITION</span>
            </div>
            """,
            unsafe_allow_html=True
        )
        st.markdown(
            """
            <div class="sharp-card" style="border-top: 4px solid #111;">
                <h1 style='margin-top:0; border-bottom:none; font-size: 2.8rem;'>智能网联汽车跨国数据合规平台 - 首页</h1>
                <p style='font-family: Lora, serif; font-size: 1.1rem; line-height: 1.6; color: #333333; margin-bottom: 0;'>
                    欢迎来到首页。此处保持空白，后续可根据需要自由添加内容。
                </p>
            </div>
            """, 
            unsafe_allow_html=True
        )
    elif st.session_state.nav_choice == "关于我们":
        st.markdown(
            """
            <div class="newsprint-masthead">
                <span>VOL. I NO. 01</span>
                <span>AUTOMOTIVE DATA COMPLIANCE REVIEW</span>
                <span>GLOBAL EDITION</span>
            </div>
            """,
            unsafe_allow_html=True
        )
        st.markdown(
            """
            <div class="sharp-card" style="border-top: 4px solid #111;">
                <h1 style='margin-top:0; border-bottom:none; font-size: 2.8rem;'>关于我们</h1>
                <p style='font-family: Lora, serif; font-size: 1.1rem; line-height: 1.6; color: #333333; margin-bottom: 0;'>
                    此处保持空白，后续可根据需要自由添加平台介绍及团队信息。
                </p>
            </div>
            """, 
            unsafe_allow_html=True
        )
    elif st.session_state.nav_choice == "案例库":
        st.markdown(
            """
            <div class="newsprint-masthead">
                <span>VOL. I NO. 01</span>
                <span>AUTOMOTIVE DATA COMPLIANCE REVIEW</span>
                <span>GLOBAL EDITION</span>
            </div>
            """,
            unsafe_allow_html=True
        )
        
        current_dir = os.path.dirname(os.path.abspath(__file__))
        case_excel_path = os.path.join(current_dir, "案例库.xlsx")
        if os.path.exists(case_excel_path):
            try:
                df_case = pd.read_excel(case_excel_path, header=None)
                cases_data = []
                for col_idx in range(1, len(df_case.columns)):
                    case_name = str(df_case.iloc[0, col_idx]).strip()
                    if not case_name or case_name == "nan":
                        continue
                    sections = {}
                    for row_idx in range(1, len(df_case)):
                        sec_title = str(df_case.iloc[row_idx, 0]).strip()
                        sec_content = str(df_case.iloc[row_idx, col_idx]).strip()
                        if sec_title and sec_title != "nan":
                            sections[sec_title] = sec_content if sec_content != "nan" else "（暂无内容）"
                    cases_data.append({
                        "case_name": case_name,
                        "sections": sections
                    })
                if st.session_state.selected_case is None:
                    st.markdown(
                        """
                        <div class="sharp-card" style="border-top: 4px solid #111;">
                            <h1 style='margin-top:0; border-bottom:none; font-size: 2.4rem;'>合规典型案例库</h1>
                            <p style='font-family: Lora, serif; font-size: 1rem; line-height: 1.5; color: #333333; margin-bottom: 0;'>
                                汇总全球数据合规与跨境执法典型案例。点击下方案例名称，即可穿透式查看包含<b>案件基本信息、基本情况、法律分析、处罚结果、合规启示与相关资料</b>在内的六大核心板块全景。
                            </p>
                        </div>
                        """, 
                        unsafe_allow_html=True
                    )
                    st.write("")
                    for i, c_item in enumerate(cases_data):
                        c_name = c_item["case_name"]
                        st.markdown(
                            f"""
                            <div class="sharp-card hard-shadow-hover" style="border-left: 6px solid #111; padding: 20px 24px; margin-bottom: 16px;">
                                <h3 style="margin-top: 0; margin-bottom: 10px; font-family: Playfair Display, serif; font-size: 1.5rem; border-bottom: none;">
                                    ⚖️ {c_name}
                                </h3>
                                <p style="font-family: Lora, serif; color: #666; margin-bottom: 15px; font-size: 0.95rem;">
                                    包含完整六维合规剖析：案件背景、事实梳理、GDPR/国内法核心条款穿透、监管逻辑、处罚裁决与出海启示。
                                </p>
                            </div>
                            """,
                            unsafe_allow_html=True
                        )
                        if st.button(f"查看完整案例报告 ➔", key=f"case_btn_{i}", use_container_width=True):
                            st.session_state.selected_case = c_name
                            st.rerun()
                else:
                    active_case = next((c for c in cases_data if c["case_name"] == st.session_state.selected_case), None)
                    if st.button("⬅ 返回案例库列表", key="back_to_cases"):
                        st.session_state.selected_case = None
                        st.rerun()
                    if active_case:
                        st.markdown(f"<h1 style='margin-top: 10px; font-size: 2.5rem;'>{active_case['case_name']}</h1>", unsafe_allow_html=True)
                        sections_order = ["案件基本信息", "案件基本情况", "法律分析", "处罚结果", "合规启示", "相关资料"]
                        for sec_title in sections_order:
                            if sec_title in active_case["sections"]:
                                content_val = active_case["sections"][sec_title]
                                st.markdown(f"<h3 style='font-family: Playfair Display, serif; margin-top: 25px; border-bottom: 1px solid #111;'>📌 {sec_title}</h3>", unsafe_allow_html=True)
                                st.markdown(f'<div class="law-content">{content_val}</div>', unsafe_allow_html=True)
                    else:
                        st.warning("未找到该案例详情。")
            except Exception as e:
                st.error(f"加载案例库表格异常: {e}")
        else:
            st.warning("未检测到 `案例库.xlsx` 文件，请确认已上传至同一目录。")
    else:
        st.markdown(
            """
            <div class="newsprint-masthead">
                <span>VOL. I NO. 01</span>
                <span>AUTOMOTIVE DATA COMPLIANCE REVIEW</span>
                <span>GLOBAL EDITION</span>
            </div>
            """,
            unsafe_allow_html=True
        )
        title_col, btn_col = st.columns([4, 1])
        with title_col:
            st.markdown(
                """
                <div class="sharp-card" style="border-top: 4px solid #111; margin-bottom: 0; height: 100%;">
                    <h1 style='margin-top:0; border-bottom:none; font-size: 2.4rem;'>智能网联汽车跨国数据合规平台</h1>
                    <p style='font-family: Lora, serif; font-size: 1rem; line-height: 1.5; color: #333333; margin-bottom: 0;'>
                        全面汇总 <b>中国、欧盟、美国</b> 三大核心司法辖区车外实景影像与关键汽车数据合规指引。<br>
                        秉持绝对清晰的网格架构与严谨审慎的编辑标准，为出境合规提供权威决策支撑。
                    </p>
                </div>
                """, 
                unsafe_allow_html=True
            )
        with btn_col:
            st.markdown(
                """
                <div class="sharp-card" style="border-top: 4px solid #111; margin-bottom: 0; display: flex; align-items: center; justify-content: center; height: 100%;">
                """,
                unsafe_allow_html=True
            )
            st.markdown('<div class="inline-term-btn">', unsafe_allow_html=True)
            if st.button("📖 术语解释总结", key="inline_terms_btn", use_container_width=True):
                st.session_state.show_terms_page = True
                st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)
        st.write("")
        if not success_db:
            st.error("主数据加载失败！请确保对应的 Excel 文件与本项目代码在同一目录下。")
        else:
            conn = sqlite3.connect(DB_FILE)
            if st.session_state.nav_choice == "法律库":
                filter_col1, filter_col2 = st.columns(2)
                with filter_col1:
                    selected_region = st.selectbox("🌐 司法辖区", ["全部", "中国", "欧盟", "美国"])
                with filter_col2:
                    if selected_region == "全部":
                        categories_df = pd.read_sql("SELECT DISTINCT category FROM compliance_laws", conn)
                    else:
                        categories_df = pd.read_sql("SELECT DISTINCT category FROM compliance_laws WHERE region = ?", conn, params=(selected_region,))
                    categories = ["全部"] + categories_df["category"].tolist()
                    selected_category = st.selectbox("📁 合规模块", categories)

                keyword = st.text_input("🔍 穿透式法规检索关键词", placeholder="如：数据出境、GDPR...")
                
                query = "SELECT region, category, law_title, sub_cat_0, sub_cat_1, content FROM compliance_laws"
                conditions = []
                params = []
                if selected_region != "全部":
                    conditions.append("region = ?")
                    params.append(selected_region)
                if selected_category != "全部":
                    conditions.append("category = ?")
                    params.append(selected_category)
                if keyword:
                    wildcard = f"%{keyword}%"
                    conditions.append("(content LIKE ? OR law_title LIKE ? OR category LIKE ? OR sub_cat_0 LIKE ? OR sub_cat_1 LIKE ?)")
                    params.extend([wildcard]*5)
                if conditions:
                    query += " WHERE " + " AND ".join(conditions)
                query += " ORDER BY region, category, sort_order"

                module_df = pd.read_sql(query, conn, params=tuple(params))

                if keyword:
                    st.markdown(f"**检索结果**：包含 <span style='background-color:#111; color:#F9F9F7; font-weight:bold; padding:2px 6px;'>“{keyword}”</span> 的内容共 **{len(module_df)}** 条", unsafe_allow_html=True)
                else:
                    st.markdown(f"**检索条件**：辖区 [{selected_region}] &nbsp;|&nbsp; 模块 [{selected_category}] &nbsp;➔&nbsp; 共计检索到 **{len(module_df)}** 条内容")

                grouped = module_df.groupby(["region", "category", "law_title"], sort=False)
                for (region_name, cat_name, law_title), group in grouped:
                    expander_label = f"📌 【{region_name}】 {law_title} ({len(group)} 条)"
                    with st.expander(expander_label, expanded=False):
                        st.markdown(f"<h4 style='font-family: Playfair Display, serif;'>{law_title}</h4>", unsafe_allow_html=True)
                        st.caption(f"归属辖区：{region_name} | 模块：{cat_name}")
                        for idx, row in group.reset_index().iterrows():
                            sc0 = row["sub_cat_0"]
                            if sc0:
                                # 展示合并后的所有适用场景标签
                                tags = [t.strip() for t in sc0.split("|") if t.strip()]
                                tags_html = "".join([f'<span class="law-tag">💡 {t}</span>' for t in tags])
                                st.markdown(tags_html, unsafe_allow_html=True)
                            content_text = row["content"]
                            if keyword:
                                content_text = content_text.replace(keyword, f"<span style='background-color:#111; color:#F9F9F7; font-weight:bold; padding:0 2px;'>{keyword}</span>")
                            st.markdown(f'<div class="law-content">{content_text}</div>', unsafe_allow_html=True)

            elif st.session_state.nav_choice == "出境全流程时间轴":
                st.markdown("### ⏱️ 数据出境全流程纵向时间轴")
                st.markdown("通过合规生命周期节点（**Phase 1：出境前准备与评估** ➔ **Phase 2：出境中实施与传输** ➔ **Phase 3：出境后合规监督**），直观展现合规实操全景。")
                all_laws_df = pd.read_sql("SELECT region, category, law_title, sub_cat_0, sub_cat_1, content FROM compliance_laws", conn)
                timeline_phases = [
                    {"title": "Phase 1：出境前准备与评估 (Data Mapping & Assessment)", 
                     "desc": "完成数据资产梳理、分类分级，执行数据出境安全评估、标准合同签署或个人信息保护认证。"},
                    {"title": "Phase 2：出境中实施与传输 (Secure Transmission & Protection)", 
                     "desc": "在符合车内处理、默认不收集、脱敏等原则下，落实跨境传输链路安全及技术保护措施。"},
                    {"title": "Phase 3：出境后合规监督 (Post-transfer Monitoring & Audit)", 
                     "desc": "建立持续合规审计机制、安全事件应急响应与境外接收方权益保障监督。"}
                ]
                phase_tabs = st.tabs([f"📌 {p['title'].split(' ')[0]} {p['title'].split(' ')[1]}" for p in timeline_phases])
                for i, p_info in enumerate(timeline_phases):
                    with phase_tabs[i]:
                        st.markdown(f"<h4 style='font-family: Playfair Display, serif;'>{p_info['title']}</h4>", unsafe_allow_html=True)
                        st.info(p_info['desc'])
                        if i == 0:
                            phase_df = all_laws_df[all_laws_df["category"].str.contains("分类|出境|安全评估|标准合同|认证", na=False) | all_laws_df["law_title"].str.contains("评估|办法|条例|规定", na=False)]
                        elif i == 1:
                            phase_df = all_laws_df[all_laws_df["content"].str.contains("传输|出境|向境外|接收|加密|安全保护", na=False)]
                            if phase_df.empty: phase_df = all_laws_df.iloc[3:7]
                        else:
                            phase_df = all_laws_df[all_laws_df["content"].str.contains("监督|审计|评估|报告|应急|处置", na=False)]
                            if phase_df.empty: phase_df = all_laws_df.iloc[7:]
                        st.markdown('<div class="timeline-container">', unsafe_allow_html=True)
                        for _, row in phase_df.iterrows():
                            region_n = row["region"]
                            law_t = row["law_title"]
                            sc0 = row["sub_cat_0"]
                            content = row["content"]
                            tag_str = f"[{region_n}] " + (sc0 if sc0 else "通用场景")
                            timeline_card_html = f"""
                            <div class="timeline-item">
                                <div class="timeline-node"></div>
                                <div class="timeline-card hard-shadow-hover">
                                    <span class="law-tag">{tag_str}</span>
                                    <h4 style="margin-top: 5px; color: #111111; font-family: Playfair Display, serif; border-bottom: none;">{law_t}</h4>
                                    <div class="law-content" style="margin-bottom: 0;">{content}</div>
                                </div>
                            </div>
                            """
                            st.markdown(timeline_card_html, unsafe_allow_html=True)
                        st.markdown('</div>', unsafe_allow_html=True)
            conn.close()
